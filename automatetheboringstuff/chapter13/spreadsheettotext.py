# Spreadsheet to text
import openpyxl, os, logging
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

def transfercelltotext(location):
    """
    Reads the content of the spreadsheet.
    """
    for folders, subfolders, filenames in os.walk(location):
        for files in filenames:
            if files.endswith(".xlsx"):
                xcel = openpyxl.load_workbook(files)
                wb = xcel.get_sheet_names()
                sheet = xcel.get_sheet_by_name(wb[0])
                logging.debug("Sheet Name: %s" % (sheet))

                content = []
                mx_row = sheet.max_row
                logging.debug("Max row: %i" % (mx_row))
                
                for col in range(1, sheet.max_column + 1):
                    for row in range(1, mx_row + 1):
                        logging.debug("i: %i" % (row))
                        if sheet.cell(row=row, column=col).value != None:
                            content.append(sheet.cell(row=row, column=col).value)
                    writetotextfile(content, col)
                    content = []

                logging.debug("Contents here: %s" % (content))


def writetotextfile(words, col):
    '''
    Writes to a textfile, gets the content and writes the
    content to a textfile depending on the column number.
    Generates a sheettotext{column}.txt.
    This is on "Write Mode" so all contents on the textfile
    will be overwritten.
    '''
    with open("sheettotext" + str(col) + ".txt", 'w') as text:
        for word in words:
            text.write(word + "\n")
    text.close()


if __name__ == "__main__":
    os.chdir('/secretfolder')
    my_dir = os.getcwd()
    logging.debug("Start of the program here.")
    transfercelltotext(my_dir)